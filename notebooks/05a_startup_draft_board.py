"""
05a_startup_draft_board.py — tiered startup-draft Excel board (tank-year build).

Consumes the star schema (fact_fantrax_adp, fact_dynasty_rankings + metrics,
dim_nfl_players, dim_position) and produces a 250-player tiered board weighted
to league scoring (SF + TE-premium + heavy IDP) and Benjamin's Year-1 tank plan.

Composite score (0-100), weights in WEIGHTS below:
  market value   35%  — KTC value + DS 3D Value+ + FP avg rank + Fantrax rank
                        + drafted% + KTC startup ADP + startup auction %
                        (mean of available; IDP has no KTC/DS, so FP + Fantrax
                        rank + drafted% carry it)
  production     20%  — trailing FP/G blended 60/40 with DS Proj 1-Yr, pct
                        within position_group
  window         25%  — positional age curve blended 70/30 with DS Proj 5-Yr
                        (cumulative pts over the 3-5yr build window)
  salary eff     20%  — FP/G per $M percentile within position_group
Tank modifiers after the blend: boost age<=25, penalize age>=28/30, plus a
Yo-Yo boost scaled by REMAINING cap-exempt runway. Yo-Yo Rule semantics: a
player is exempt until his 20th career NFL game is PLAYED — runway burns per
game played, never by calendar. Exact counts via nflverse weekly stats for
entry_year >= GAMES_CHECK_FLOOR.

Tiers band by composite rank (1a..4), with age guardrails from the tier
definitions (1a <=26, 1b <=27, 2a <=28) demoting violators one band.

Judgment overlay: data/raw/draft_board_notes.csv (gsis_id, tier_override,
arc_note, strategic_note). Bespoke notes cover tiers 1a-2b; everyone else gets
rule-generated flag strings. Missing file -> script writes a prefilled template.

Output: data/outputs/startup_draft_board.xlsx
  "Offense" / "Defense" sheets — 250 each, per-side tier ladders + side Rank,
      with cross-side Score and Ovr Rank kept for comparison at the pick.
      Defense drops the offense-only metric columns (ADP/KTC/DS projections).
  "Data" sheet — full scored pool, flat, pivot/slicer-ready.

Run from repo root or notebooks/:  python notebooks/05a_startup_draft_board.py
"""
import sys
from pathlib import Path

for _p in (Path.cwd() / "notebooks", Path.cwd(), Path(__file__).resolve().parent):
    if (_p / "etl_helpers.py").exists():
        sys.path.insert(0, str(_p))
        break

import numpy as np
import pandas as pd
from etl_helpers import CFG, DATA

# ── Config ───────────────────────────────────────────────────────────
BOARD_SIZE = 250                       # per side: Offense board 250, Defense board 250
WEIGHTS = {"market": 0.35, "production": 0.20, "age_arc": 0.25, "salary_eff": 0.20}
# Tank modifiers (points on the 0-100 composite)
BOOST_YOUNG, YOUNG_AGE_MAX = 5.0, 25
PENALTY_VET, VET_AGE_MIN = -7.0, 28
PENALTY_OLD, OLD_AGE_MIN = -3.0, 30   # stacks on PENALTY_VET
BOOST_ML_ELIGIBLE = 3.0                # scaled by remaining runway (games_left/20)
ML_GAMES_LIMIT = 20                    # Yo-Yo Rule: exempt until the 20th career NFL
                                       # game is PLAYED — runway burns per game played,
                                       # not per season stashed
GAMES_CHECK_FLOOR = 2022               # entry_year >= this gets an exact games count
PROD_PROJ_BLEND = 0.40                 # production = 60% trailing FP/G + 40% Proj 1-Yr
WINDOW_PROJ_BLEND = 0.30               # window = 70% age curve + 30% Proj 5-Yr pct
# Tier bands: (max composite rank, tier, max age or None). Violators demote a band.
TIER_BANDS = [(8, "1a", 26), (25, "1b", 27), (60, "2a", 28),
              (100, "2b", None), (175, "3", None), (BOARD_SIZE, "4", None)]
# Positional age-curve offsets (effective age = age - offset)
AGE_OFFSET = {"QB": 3, "TE": 1, "RB": -1}
ADP_OUTLIER_GAP = 40                   # composite rank vs ADP rank review flag

OUT_DIR = DATA / "outputs"
NOTES_CSV = DATA / "raw" / "draft_board_notes.csv"
BOARD_XLSX = OUT_DIR / "startup_draft_board.xlsx"

IDP_GROUPS = {"DL", "LB", "DB"}


# ── Load + join ──────────────────────────────────────────────────────
def load_pool() -> pd.DataFrame:
    adp = pd.read_parquet(CFG.path(CFG.fact_name))
    adp = adp[adp["capture_date"] == adp["capture_date"].max()].copy()

    pos = pd.read_parquet(DATA / "dim_position.parquet")
    pos_map = pos.set_index("position_raw")[["position_group", "side_of_ball"]]
    # Multi-position rows ("DL,LB") fall back to their first listed position.
    first_pos = adp["position_raw"].str.split(",").str[0]
    lookup = adp["position_raw"].where(adp["position_raw"].isin(pos_map.index), first_pos)
    adp["position_group"] = lookup.map(pos_map["position_group"])
    adp["side_of_ball"] = lookup.map(pos_map["side_of_ball"])

    npl = pd.read_parquet(CFG.path(CFG.nfl_players_name))
    npl = npl[["gsis_id", "entry_year", "draft_year", "draft_round", "birth_date"]]
    npl = npl.dropna(subset=["gsis_id"])
    pool = adp.merge(npl, on="gsis_id", how="left")
    # Undrafted vets have NaN draft_year; entry_year always populated for matched ids.
    pool["draft_year"] = pool["draft_year"].fillna(pool["entry_year"])
    pool["years_in_league"] = (CFG.draft_year - pool["entry_year"]).clip(lower=0)
    # Fantrax omits age for some players — backfill from registry birth_date.
    bday = pd.to_datetime(pool["birth_date"], errors="coerce")
    reg_age = (pd.Timestamp.today() - bday).dt.days / 365.25
    pool["age"] = pool["age"].fillna(reg_age.round(1))
    return pool


# metric_key -> board column. Each metric_key owns exactly ONE source (the
# etl.SOURCE_PREFIX invariant — source lives on dim_dynasty_metric), so the source
# is derivable and not needed here. This map is pure board presentation: which
# metrics the draft board surfaces and what it calls them, nothing about the registry.
METRIC_MAP = {
    "value": "ktc_value",
    "startup_adp": "startup_adp",
    "startup_avg_auction_pct": "startup_auction_pct",
    "std_liquidity": "liquidity",
    "ds_value": "ds_value",
    "proj_1yr": "proj_1yr",
    "proj_3yr": "proj_3yr",
    "proj_5yr": "proj_5yr",
    "proj_10yr": "proj_10yr",
    "avg": "fp_avg",
    "best": "fp_best",
    "worst": "fp_worst",
    "stddev": "fp_stddev",
}


def load_dynasty_metrics() -> pd.DataFrame:
    """Per-gsis dynasty signals (latest SF snapshot per source): values, startup
    market, FP consensus spread, DS multi-year projections (cumulative pts)."""
    # The metrics EAV carries gsis_id directly (the old backbone fact is gone), so
    # key off it straight away — no separate rankings table to join through.
    m = pd.read_parquet(DATA / "fact_dynasty_ranking_metrics.parquet")
    m = m[m["format"] == "SF"].dropna(subset=["gsis_id"])
    m = m[m["snapshot_date"] == m.groupby("source_name")["snapshot_date"].transform("max")]

    cols = []
    for key, col in METRIC_MAP.items():
        # metric_key is 1:1 with source, so filtering on it alone selects the right
        # source's rows (the latest-snapshot-per-source filter above already applied).
        s = m[m["metric_key"] == key]
        cols.append(s.groupby("gsis_id")["metric_num"].max().rename(col))
    return pd.concat(cols, axis=1).reset_index()


def load_drafted() -> dict:
    """gsis_id -> team_key for players already taken in the live startup draft,
    read from the derived current roster (`fact_fantasy_teams`, built by 02e from
    the `fact_roster_transactions` ledger). Returns {} if the ledger hasn't been
    built yet, so the board still runs pre-draft. Re-run 04w -> 02d -> 02e between
    picks to refresh; this surfaces who's gone and to whom on the board."""
    p = DATA / "fact_fantasy_teams.parquet"
    if not p.exists():
        return {}
    fft = pd.read_parquet(p).dropna(subset=["gsis_id"])
    return dict(zip(fft["gsis_id"], fft["team_key"]))


def load_career_games(pool: pd.DataFrame) -> pd.Series:
    """Exact NFL games played (REG+POST) per gsis_id for recent entrants.
    entry_year < GAMES_CHECK_FLOOR is assumed >= ML_GAMES_LIMIT games."""
    import nflreadpy as nfl
    recent = pool[pool["entry_year"] >= GAMES_CHECK_FLOOR]["gsis_id"].dropna().unique()
    seasons = list(range(GAMES_CHECK_FLOOR, CFG.draft_year))
    stats = nfl.load_player_stats(seasons=seasons).to_pandas()
    id_col = "player_id" if "player_id" in stats.columns else "gsis_id"
    stats = stats[stats[id_col].isin(recent)]
    games = stats.groupby(id_col).size().rename("career_games")
    games.index.name = "gsis_id"
    return games


# ── Scoring ──────────────────────────────────────────────────────────
def pct(s: pd.Series) -> pd.Series:
    return s.rank(pct=True) * 100


def score(pool: pd.DataFrame) -> pd.DataFrame:
    df = pool.copy()

    # Market: mean of available percentile signals. Rank metrics invert (low=good).
    sig = pd.DataFrame(index=df.index)
    sig["ktc"] = pct(df["ktc_value"])
    sig["ds"] = pct(df["ds_value"])
    sig["fp"] = pct(-df["fp_avg"])
    sig["fx"] = pct(-df["overall_rank"])
    sig["px"] = pct(df["percent_drafted"])      # cross-league demand; covers IDP
    sig["sadp"] = pct(-df["startup_adp"])       # actual startup market price
    sig["auc"] = pct(df["startup_auction_pct"])
    df["market_score"] = sig.mean(axis=1, skipna=True)

    # Production: trailing FP/G blended with DS Proj 1-Yr (both pct within
    # position_group; projection covers offense top-250 only). Rookies without
    # either signal sit neutral (50) — youth boost handles the upside.
    grp = df.groupby("position_group")
    trailing = grp["fpts_per_game"].transform(pct)
    proj1 = grp["proj_1yr"].transform(pct)
    w_proj = np.where(proj1.notna(), PROD_PROJ_BLEND, 0.0)
    df["production_score"] = (trailing.fillna(50) * (1 - w_proj)
                              + proj1.fillna(0) * w_proj)
    df["fpg_per_m"] = df["fpts_per_game"] / (df["salary"] / 1e6)
    df["salary_eff_score"] = grp["fpg_per_m"].transform(pct)
    no_sample = df["is_rookie"] & (df["fpts_per_game"].isna() | (df["fpts_per_game"] <= 0))
    df.loc[no_sample & df["proj_1yr"].isna(), "production_score"] = 50.0
    df.loc[no_sample, "salary_eff_score"] = 50.0

    # Window: positional age curve blended with DS Proj 5-Yr (cumulative pts
    # over exactly the 3-5yr window the build targets; offense only).
    eff_age = df["age"] - df["position_group"].map(AGE_OFFSET).fillna(0)
    bins = [-np.inf, 23, 25, 26, 27, 28, 29, np.inf]
    vals = [100, 90, 75, 60, 45, 30, 15]
    age_curve = pd.cut(eff_age, bins=bins, labels=vals).astype(float).fillna(50)
    proj5 = grp["proj_5yr"].transform(pct)
    w_p5 = np.where(proj5.notna(), WINDOW_PROJ_BLEND, 0.0)
    df["age_arc_score"] = age_curve * (1 - w_p5) + proj5.fillna(0) * w_p5

    # Projection trajectory: avg yr-3 pace vs yr-1 (ascending > 0 > declining).
    df["proj_traj"] = (df["proj_3yr"] / 3) / df["proj_1yr"] - 1

    df["composite"] = (
        WEIGHTS["market"] * df["market_score"].fillna(50)
        + WEIGHTS["production"] * df["production_score"].fillna(50)
        + WEIGHTS["age_arc"] * df["age_arc_score"]
        + WEIGHTS["salary_eff"] * df["salary_eff_score"].fillna(50)
    )
    df["composite"] += np.where(df["age"] <= YOUNG_AGE_MAX, BOOST_YOUNG, 0)
    df["composite"] += np.where(df["age"] >= VET_AGE_MIN, PENALTY_VET, 0)
    df["composite"] += np.where(df["age"] >= OLD_AGE_MIN, PENALTY_OLD, 0)
    # Yo-Yo boost scales with remaining cap-exempt runway, not a binary flag.
    df["composite"] += BOOST_ML_ELIGIBLE * (df["ml_games_left"] / ML_GAMES_LIMIT)
    df["composite"] = df["composite"].clip(0, 100).round(1)

    df = df.sort_values("composite", ascending=False).reset_index(drop=True)
    df["board_rank"] = df.index + 1                      # cross-side (overall)
    df["side_rank"] = df.groupby("side_of_ball")["composite"] \
                        .rank(ascending=False, method="first")

    # ADP sanity flag (offense only — IDP has no Fantrax ADP).
    adp_rank = df["adp"].rank()
    df["adp_outlier"] = (df["board_rank"] - adp_rank).abs() > ADP_OUTLIER_GAP
    return df


def assign_tiers(df: pd.DataFrame) -> pd.DataFrame:
    """Per-side tier ladders: 1a-4 bands applied to side_rank within each of
    Offense and Defense, with the tier definitions' age guardrails."""
    tiers = ["" for _ in range(len(df))]
    order = [b[1] for b in TIER_BANDS]
    on_board = df["side_of_ball"].isin(["Offense", "Defense"]) & (df["side_rank"] <= BOARD_SIZE)
    for i in df.index[on_board]:
        rank = df["side_rank"].iat[i]
        for max_rank, tier, age_max in TIER_BANDS:
            if rank <= max_rank:
                age = df["age"].iat[i]
                if age_max is not None and pd.notna(age) and age > age_max:
                    tier = order[min(order.index(tier) + 1, len(order) - 1)]
                tiers[i] = tier
                break
    df["tier"] = tiers
    return df


# ── Notes overlay ────────────────────────────────────────────────────
def rule_notes(row) -> tuple[str, str]:
    """Rule-generated Arc / Strategic core for players without bespoke notes."""
    age = row["age"]
    arc = ("Ascending" if pd.notna(age) and age <= 25 else
           "Prime window" if pd.notna(age) and age <= 27 else "Declining")
    if pd.notna(age):
        arc += f" (age {age:.0f})"
    traj = row.get("proj_traj")
    if pd.notna(traj) and abs(traj) >= 0.08:
        arc += "; proj " + (f"+{traj:.0%} by yr-3" if traj > 0 else f"{traj:.0%} by yr-3")
    flags = []
    if pd.notna(age) and age >= VET_AGE_MIN and row.get("production_score", 0) >= 70:
        flags.append("Sell-high vet / rental")
    if row.get("salary_eff_score", 0) >= 75 and row.get("market_score", 100) <= 50:
        flags.append("Buy-low value")
    if pd.notna(row.get("salary")):
        flags.append(f"Drop hit ${row['salary'] * 0.5 / 1e6:.1f}M (Yr-1 50%)")
    return arc, "; ".join(flags)


def metric_suffix(row, liq_lo, liq_hi, sd_hi) -> str:
    """Data callouts appended to every board note (bespoke and rule-generated):
    Yo-Yo runway, 5-yr projection window, expert consensus spread, KTC market
    liquidity, startup-market price gap."""
    bits = []
    gl = row.get("ml_games_left")
    if pd.notna(gl) and gl > 0:
        bits.append(f"Yo-Yo: {gl:.0f}/{ML_GAMES_LIMIT} exempt games left"
                    + (" (full runway)" if gl >= ML_GAMES_LIMIT else ""))
    elif row.get("entry_year", 0) >= GAMES_CHECK_FLOOR:
        bits.append("Yo-Yo window closed")
    if pd.notna(row.get("proj_5yr")):
        bits.append(f"proj {row['proj_5yr']:,.0f} pts/5yr")
    sd = row.get("fp_stddev")
    if pd.notna(sd):
        if sd >= sd_hi:
            bits.append(f"experts split (FP {row['fp_best']:.0f}-{row['fp_worst']:.0f}, "
                        f"σ{sd:.0f})")
    liq = row.get("liquidity")
    if pd.notna(liq):
        if liq >= liq_hi:
            bits.append("liquid market")
        elif liq <= liq_lo:
            bits.append("thin market — price patiently")
    sadp = row.get("startup_adp")
    if pd.notna(sadp):
        gap = sadp - row["board_rank"]
        if gap >= ADP_OUTLIER_GAP:
            bits.append(f"startup market sleeping (SADP {sadp:.0f})")
        elif gap <= -ADP_OUTLIER_GAP:
            bits.append(f"market price rich (SADP {sadp:.0f})")
    return " | ".join(bits)


def apply_notes(df: pd.DataFrame) -> pd.DataFrame:
    gen = df.apply(rule_notes, axis=1, result_type="expand")
    df["arc_note"], df["strategic_note"] = gen[0], gen[1]
    if NOTES_CSV.exists():
        notes = pd.read_csv(NOTES_CSV, dtype={"gsis_id": str})
        notes = notes.dropna(subset=["gsis_id"]).drop_duplicates("gsis_id").set_index("gsis_id")
        for col in ("arc_note", "strategic_note"):
            if col in notes:
                ov = df["gsis_id"].map(notes[col])
                df[col] = ov.where(ov.notna() & (ov.astype(str).str.strip() != ""), df[col])
        if "tier_override" in notes:
            ov = df["gsis_id"].map(notes["tier_override"])
            df["tier"] = ov.where(ov.notna() & (ov.astype(str).str.strip() != ""), df["tier"])
    else:
        tmpl = df.head(BOARD_SIZE)[["gsis_id", "player_name", "position_raw", "tier"]].copy()
        tmpl[["tier_override", "arc_note", "strategic_note"]] = ""
        NOTES_CSV.parent.mkdir(parents=True, exist_ok=True)
        tmpl.to_csv(NOTES_CSV, index=False)
        print(f"Notes template created -> {NOTES_CSV} (fill arc/strategic notes, rerun)")

    # Metric callouts append to every board-row note, bespoke included.
    liq_lo, liq_hi = df["liquidity"].quantile([0.30, 0.70])
    sd_hi = df["fp_stddev"].quantile(0.75)
    on_board = df["side_of_ball"].isin(["Offense", "Defense"]) & (df["side_rank"] <= BOARD_SIZE)
    suffix = df[on_board].apply(metric_suffix, axis=1, args=(liq_lo, liq_hi, sd_hi))
    has_sfx = suffix.str.len() > 0
    df.loc[on_board, "strategic_note"] = (
        df.loc[on_board, "strategic_note"].str.rstrip()
        + np.where(has_sfx, " ‖ " + suffix, "")
    )
    return df


# ── Excel output ─────────────────────────────────────────────────────
BOARD_COLS = {
    "tier": "Tier", "side_rank": "Rank", "board_rank": "Ovr Rank",
    "drafted_by": "Drafted By",
    "player_name": "Player", "position_raw": "Position",
    "position_group": "Pos Group", "nfl_team": "NFL Team", "age": "Age",
    "draft_year": "Draft Year", "years_in_league": "Years In League",
    "overall_rank": "Fantrax Rank", "adp": "ADP", "startup_adp": "Startup ADP",
    "salary": "Salary", "fpts_per_game": "FP/G", "percent_drafted": "Ros%",
    "ktc_value": "KTC Value", "proj_1yr": "Proj 1-Yr", "proj_5yr": "Proj 5-Yr",
    "proj_10yr": "Proj 10-Yr",
    "fp_stddev": "FP StdDev", "liquidity": "Liquidity",
    "ml_games_left": "ML Games Left", "composite": "Score",
    "arc_note": "Arc Note", "strategic_note": "Strategic Note",
}
# IDP has no Fantrax ADP / KTC / DS coverage — drop those columns on Defense.
DEFENSE_DROP = ["adp", "startup_adp", "ktc_value", "proj_1yr", "proj_5yr",
                "proj_10yr", "liquidity"]
# Data sheet adds the rest of the metric block + pillar scores for pivots.
DATA_EXTRA_COLS = {
    "side_of_ball": "Side of Ball", "ds_value": "DS 3D Value+",
    "proj_3yr": "Proj 3-Yr",
    "proj_traj": "Proj Traj", "fp_avg": "FP Avg Rank", "fp_best": "FP Best Rank",
    "fp_worst": "FP Worst Rank", "startup_auction_pct": "Startup Auction %",
    "career_games": "Career Games", "ml_eligible": "ML Eligible",
    "adp_outlier": "ADP Outlier", "market_score": "Market Score",
    "production_score": "Production Score", "age_arc_score": "Window Score",
    "salary_eff_score": "Salary Eff Score",
}
TIER_FILL = {"1a": "FFD966", "1b": "FFE699", "2a": "A9D08E", "2b": "C6E0B4",
             "3": "BDD7EE", "4": "D9D9D9"}


NUM_FMT = {"Salary": '"$"#,##0.0,,"M"', "ADP": "0.0", "FP/G": "0.0",
           "Ros%": '0.0"%"', "Age": "0", "Score": "0.0",
           "Draft Year": "0", "Years In League": "0", "Fantrax Rank": "0",
           "Rank": "0", "Ovr Rank": "0",
           "Startup ADP": "0.0", "KTC Value": "#,##0", "Proj 1-Yr": "#,##0",
           "Proj 5-Yr": "#,##0", "Proj 10-Yr": "#,##0",
           "FP StdDev": "0.0", "Liquidity": "0.0",
           "ML Games Left": "0"}
COL_WIDTHS = {"Player": 24, "Position": 9, "Pos Group": 10,
              "Arc Note": 38, "Strategic Note": 80, "NFL Team": 9}


def write_board_sheet(wb, name: str, board: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    ws = wb.create_sheet(name)
    ws.append(list(board.columns))
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F3864")
        c.alignment = Alignment(horizontal="center")
    for row in board.itertuples(index=False):
        ws.append(list(row))

    n_rows, n_cols = len(board) + 1, len(board.columns)
    table = Table(displayName=f"{name}Board",
                  ref=f"A1:{get_column_letter(n_cols)}{n_rows}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(table)
    ws.freeze_panes = "D2"

    tier_col = list(board.columns).index("Tier") + 1
    for r in range(2, n_rows + 1):
        fill_hex = TIER_FILL.get(ws.cell(r, tier_col).value)
        if fill_hex:
            fill = PatternFill("solid", fgColor=fill_hex)
            for c in range(1, n_cols + 1):
                ws.cell(r, c).fill = fill

    for idx, col in enumerate(board.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(col, 11)
        if col in NUM_FMT:
            for r in range(2, n_rows + 1):
                ws.cell(r, idx).number_format = NUM_FMT[col]


def write_excel(df: pd.DataFrame) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    boards = {}
    for side in ("Offense", "Defense"):
        cols = {k: v for k, v in BOARD_COLS.items()
                if not (side == "Defense" and k in DEFENSE_DROP)}
        side_df = df[(df["side_of_ball"] == side) & (df["side_rank"] <= BOARD_SIZE)]
        boards[side] = side_df.sort_values("side_rank")[list(cols)].rename(columns=cols)
        write_board_sheet(wb, side, boards[side])

    full_cols = {**BOARD_COLS, **DATA_EXTRA_COLS}
    full = df[list(full_cols)].rename(columns=full_cols)
    ws2 = wb.create_sheet("Data")
    ws2.append(list(full.columns))
    for row in full.itertuples(index=False):
        ws2.append(list(row))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(full.columns))}{len(full) + 1}"
    ws2.freeze_panes = "A2"

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(BOARD_XLSX)
    print(f"Board written -> {BOARD_XLSX} "
          f"(Offense {len(boards['Offense'])}, Defense {len(boards['Defense'])}, "
          f"Data {len(full)})")


# ── Main ─────────────────────────────────────────────────────────────
def main() -> pd.DataFrame:
    pool = load_pool()
    pool = pool.merge(load_dynasty_metrics(), on="gsis_id", how="left")

    games = load_career_games(pool)
    pool = pool.merge(games, on="gsis_id", how="left")
    recent = pool["entry_year"] >= GAMES_CHECK_FLOOR
    pool["career_games"] = pool["career_games"].where(~recent | pool["career_games"].notna(), 0)
    # Yo-Yo runway: exempt games remaining before the 20th career game is played.
    # Pre-floor entrants are assumed past the limit (runway 0).
    pool["ml_games_left"] = np.where(
        recent, (ML_GAMES_LIMIT - pool["career_games"]).clip(lower=0), 0)
    pool["ml_eligible"] = pool["ml_games_left"] > 0

    # Live-draft availability: flag players already taken (and by whom). Non-
    # destructive — kept on the board so runs stay visible; filter in Excel.
    drafted = load_drafted()
    pool["drafted_by"] = pool["gsis_id"].map(drafted)
    pool["available"] = pool["drafted_by"].isna()
    print(f"[info] live draft: {len(drafted)} players taken, "
          f"{int(pool['available'].sum())} of {len(pool)} pool still available")

    df = assign_tiers(score(pool))
    df = apply_notes(df)
    write_excel(df)

    for side in ("Offense", "Defense"):
        sd = df[(df["side_of_ball"] == side) & (df["tier"] != "")]
        print(f"\n{side} tier counts:")
        print(sd["tier"].value_counts().sort_index().to_string())
        cols = ["side_rank", "tier", "player_name", "position_raw", "age", "salary",
                "composite"]
        print(f"{side} top 10:")
        print(sd.nsmallest(10, "side_rank")[cols].to_string(index=False))
    return df


if __name__ == "__main__":
    main()
